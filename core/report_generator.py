from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from core.validator import ValidationResult


class ReportGenerator:
    def _error_frame(self, df: pd.DataFrame, validation_result: ValidationResult) -> pd.DataFrame:
        row_errors: Dict[int, List] = defaultdict(list)
        for err in validation_result.errors:
            row_errors[err.row_number - 1].append(err)
        invalid_df = df.loc[sorted(row_errors.keys())].copy() if row_errors else df.iloc[0:0].copy()
        invalid_df["error_count"] = invalid_df.index.map(lambda i: len(row_errors.get(i, [])))
        invalid_df["error_types"] = invalid_df.index.map(
            lambda i: ", ".join(sorted({e.error_type for e in row_errors.get(i, [])}))
        )
        invalid_df["error_details"] = invalid_df.index.map(
            lambda i: " | ".join([e.error_message for e in row_errors.get(i, [])])
        )
        return invalid_df

    def generate_full_report(self, df: pd.DataFrame, validation_result: ValidationResult, score_result: dict) -> bytes:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            wb = writer.book
            fmt_header = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#1F2937", "border": 1})
            fmt_green = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#10B981", "border": 1})
            fmt_red = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#EF4444", "border": 1})
            fmt_border = wb.add_format({"border": 1})
            summary_rows = [
                ("Report Generated", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
                ("Source File", "Uploaded Dataset"),
                ("Total Rows Processed", validation_result.total_rows),
                ("Valid Rows", len(validation_result.valid_row_indices)),
                ("Invalid Rows", len(validation_result.invalid_row_indices)),
                ("Readiness Score", f"{score_result['score']}%"),
                ("Score Band", score_result["band_label"]),
            ]
            pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(writer, sheet_name="Executive Summary", index=False, startrow=3)
            ws1 = writer.sheets["Executive Summary"]
            ws1.merge_range("A1:F1", "DataForge — Purify. Validate. Migrate.", wb.add_format({"bold": True, "font_size": 18}))
            ws1.merge_range("A2:F2", "Validation Report", wb.add_format({"bold": True, "font_size": 12}))
            ws1.write("A12", "Top 3 Critical Issues", fmt_header)
            top_issues = sorted(validation_result.error_counts_by_type.items(), key=lambda x: x[1], reverse=True)[:3]
            for i, (k, v) in enumerate(top_issues, start=13):
                ws1.write(f"A{i}", k, fmt_border)
                ws1.write(f"B{i}", v, fmt_border)
            valid_df = df.iloc[validation_result.valid_row_indices].copy()
            valid_df.to_excel(writer, sheet_name="Migration Ready Data", index=False)
            ws2 = writer.sheets["Migration Ready Data"]
            for c in range(len(valid_df.columns)):
                ws2.write(0, c, valid_df.columns[c], fmt_green)
            ws2.write(len(valid_df) + 2, 0, f"Total migration-ready rows: {len(valid_df)}", fmt_border)
            invalid_df = self._error_frame(df, validation_result)
            invalid_df.to_excel(writer, sheet_name="Invalid Data + Error Log", index=False)
            ws3 = writer.sheets["Invalid Data + Error Log"]
            for c in range(len(invalid_df.columns)):
                ws3.write(0, c, invalid_df.columns[c], fmt_red)
            pivot = pd.DataFrame(0, index=df.columns.tolist(), columns=["NULL", "TYPE", "FORMAT", "RANGE", "DUPLICATE", "REFERENCE"])
            for e in validation_result.errors:
                if e.column_name in pivot.index and e.error_type in pivot.columns:
                    pivot.loc[e.column_name, e.error_type] += 1
            pivot.to_excel(writer, sheet_name="Error Summary by Column")
            ws4 = writer.sheets["Error Summary by Column"]
            ws4.set_column(0, len(pivot.columns), 18)
            ws4.conditional_format(1, 1, len(pivot.index), len(pivot.columns), {"type": "3_color_scale", "min_color": "#FFFFFF", "mid_color": "#FCA5A5", "max_color": "#7F1D1D"})
            recs = []
            effort_map = {
                "NULL": "Low — fill missing values with default or lookup",
                "FORMAT": "Low — standardize format with regex replace",
                "RANGE": "Medium — review and correct out-of-range values",
                "TYPE": "Medium — convert column to correct data type",
                "DUPLICATE": "High — manual review and deduplication needed",
                "REFERENCE": "High — cross-reference with master data table",
            }
            grouped = defaultdict(int)
            for e in validation_result.errors:
                grouped[(e.column_name, e.error_type)] += 1
            for idx, ((col, typ), cnt) in enumerate(sorted(grouped.items(), key=lambda x: x[1], reverse=True), start=1):
                fix_text = {
                    "NULL": f"Fill in missing {col} value",
                    "FORMAT": f"Correct format in {col}",
                    "RANGE": f"Review {col} values against allowed range",
                    "TYPE": f"Convert {col} to expected type",
                    "DUPLICATE": f"Remove or merge duplicate records for {col}",
                    "REFERENCE": f"Assign valid master key values for {col}",
                }.get(typ, f"Fix data quality issues in {col}")
                recs.append([idx, col, typ, cnt, effort_map.get(typ, "Medium"), fix_text])
            pd.DataFrame(
                recs, columns=["Priority", "Column Name", "Error Type", "Error Count", "Estimated Effort", "Fix Instruction"]
            ).to_excel(writer, sheet_name="Fix Recommendations", index=False, startrow=1)
            ws5 = writer.sheets["Fix Recommendations"]
            ws5.write("A1", "Prioritized Fix Recommendations", fmt_header)
            for sheet in [ws1, ws2, ws3, ws4, ws5]:
                sheet.freeze_panes(1, 0)
                sheet.set_column(0, 40, 15)
        return buffer.getvalue()

    def generate_valid_data_export(self, df: pd.DataFrame, valid_indices: List[int]) -> bytes:
        buffer = BytesIO()
        valid_df = df.iloc[valid_indices].copy()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            valid_df.to_excel(writer, sheet_name="Valid Data", index=False)
            wb = writer.book
            ws = writer.sheets["Valid Data"]
            header_fmt = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#10B981", "border": 1})
            for col, name in enumerate(valid_df.columns):
                ws.write(0, col, name, header_fmt)
            ws.freeze_panes(1, 0)
            ws.set_column(0, len(valid_df.columns) - 1, 15)
        return buffer.getvalue()

    def generate_error_report(self, df: pd.DataFrame, validation_result: ValidationResult) -> bytes:
        base = self._error_frame(df, validation_result)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            base.to_excel(writer, sheet_name="Error Report", index=False)
            wb = writer.book
            ws = writer.sheets["Error Report"]
            header_fmt = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#EF4444", "border": 1})
            for col, name in enumerate(base.columns):
                ws.write(0, col, name, header_fmt)
            ws.set_column(0, len(base.columns) - 1, 15)
            ws.freeze_panes(1, 0)
        buffer.seek(0)
        wb2 = load_workbook(buffer)
        ws2 = wb2["Error Report"]
        red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        blue = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        col_pos = {name: i + 1 for i, name in enumerate(base.columns)}
        for err in validation_result.errors:
            if err.row_number - 1 not in base.index or err.column_name not in col_pos:
                continue
            excel_row = list(base.index).index(err.row_number - 1) + 2
            cell = ws2.cell(row=excel_row, column=col_pos[err.column_name])
            cell.fill = red if err.severity == "Critical" else yellow if err.severity == "Warning" else blue
        out = BytesIO()
        wb2.save(out)
        return out.getvalue()
